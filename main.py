import pandas as pd
import json
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

with open('Installments.json', 'r', encoding='utf-8') as file:
    data = json.load(file)

df_main = pd.json_normalize(data)

column_renaming = {...}

df_main.rename(columns=column_renaming, inplace=True)

columns_to_drop = [...]

df_main = df_main.drop(columns=[col for col in columns_to_drop if col in df_main.columns])

messenger_mapping = {...}

df_main['Internal Messenger Type'] = df_main['Internal Messenger Type'].replace(messenger_mapping)

df_main['Status'] = df_main['Status'].str.capitalize()

df_main['Gender'] = df_main['Gender'].str.capitalize()

output_file = 'output.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_main.to_excel(writer, sheet_name='Main', index=False)

wb = openpyxl.load_workbook(output_file)

font = Font(name='Vazirmatn')
alignment = Alignment(horizontal='center', vertical='center')
light_yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
light_green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

sheet = wb['Main']
for row in sheet:
    for cell in row:
        cell.font = font
        cell.alignment = alignment
        cell.number_format = '@'
        if cell.row == 1:
            cell.fill = light_yellow_fill
        else:
            cell.fill = light_green_fill

    for column_cells in sheet.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length

wb.save(output_file)
